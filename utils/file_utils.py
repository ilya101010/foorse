from pathlib import Path
import os
import xlrd
import openpyxl
import pandas as pd
from openpyxl.utils import get_column_letter

def get_html_files(directory):
    return list(Path(directory).rglob("*.html"))

def get_excel_files(directory):
    # Use rglob to find both .xlsx and .xls files
    return list(Path(directory).rglob("*.xlsx")) + list(Path(directory).rglob("*.xls"))

def pick_reference_file(html_files):
    return max(html_files, key=os.path.getsize)

def read_xls_to_dataframe(file_path):
    # Open the workbook
    workbook = xlrd.open_workbook(file_path, formatting_info=True)
    
    result = []
    
    # Iterate through each sheet in the workbook
    for sheet_index in range(workbook.nsheets):
        sheet = workbook.sheet_by_index(sheet_index)
        sheet_name = sheet.name
        
        # Read the data into a pandas DataFrame
        data = []
        for row_index in range(sheet.nrows):
            row_values = []
            for col_index in range(sheet.ncols):
                cell_value = sheet.cell_value(row_index, col_index)
                row_values.append(cell_value)
            data.append(row_values)
        
        df = pd.DataFrame(data)
        
        # Handle merged cells
        for merge in sheet.merged_cells:
            r1, r2, c1, c2 = merge
            merge_value = sheet.cell_value(r1, c1)
            df.iloc[r1:r2, c1:c2] = merge_value
        
        # Remove hidden rows and columns
        hidden_rows = set()
        hidden_cols = set()
        
        for row_index in range(sheet.nrows):
            if sheet.rowinfo_map.get(row_index, None) and sheet.rowinfo_map[row_index].hidden:
                hidden_rows.add(row_index)
        
        for col_index in range(sheet.ncols):
            if sheet.colinfo_map.get(col_index, None) and sheet.colinfo_map[col_index].hidden:
                hidden_cols.add(col_index)
        
        df = df.drop(hidden_rows, axis=0).drop(hidden_cols, axis=1)
        
        # Append the sheet name and DataFrame as a tuple to the result list
        result.append((sheet_name, df))
    
    return result

def read_xlsx_to_dataframe(file_path):
    # Open the workbook
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    
    result = []
    
    # Iterate through each sheet in the workbook
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        
        # Read the data into a list of lists
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(list(row))
        
        # Create a DataFrame with object dtype to handle mixed data types
        df = pd.DataFrame(data, dtype=object)
        
        # Handle merged cells
        for merge in sheet.merged_cells.ranges:
            merge_value = sheet.cell(merge.min_row, merge.min_col).value
            df.iloc[merge.min_row-1:merge.max_row, merge.min_col-1:merge.max_col] = merge_value
        
        # Remove hidden rows
        hidden_rows = set()
        for row_index in range(1, sheet.max_row + 1):
            if sheet.row_dimensions[row_index].hidden:
                hidden_rows.add(row_index - 1)
        df = df.drop(hidden_rows, axis=0)
        
        # Detect hidden columns
        max_col = sheet.max_column
        cols = [get_column_letter(i) for i in range(1, max_col + 1)]
        hidden_cols = []
        last_hidden = 0
        for i, col in enumerate(cols):
            if sheet.column_dimensions[col].hidden:
                hidden_cols.append(i)
                last_hidden = sheet.column_dimensions[col].max
            elif i + 1 <= last_hidden:
                hidden_cols.append(i)
        
        df = df.drop(hidden_cols, axis=1)
        
        # Append the sheet name and DataFrame as a tuple to the result list
        result.append((sheet_name, df))
    
    return result